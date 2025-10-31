from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from datetime import timedelta, datetime, date
from django.db.models import Count, Q, Sum, F, Value
from django.db.models.functions import Coalesce, Concat
from django.db.models import Value
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django import forms
import csv
import json
from django.core.serializers.json import DjangoJSONEncoder
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

from clientes.models import Cliente
from mascotas.models import Mascota
from citas.models import Cita
from tienda.models import Producto, Categoria, Orden
from django.contrib.auth.models import User
from notificaciones.models import Notificacion, Recordatorio

def staff_required(login_url=None):
    return user_passes_test(lambda u: u.is_staff, login_url=login_url)

def veterinario_required(login_url=None):
    return user_passes_test(lambda u: u.groups.filter(name='Veterinarios').exists() or u.is_staff, login_url=login_url)

def dashboard(request):
    hoy = timezone.now().date()

    # ESTADÍSTICAS EXISTENTES
    total_clientes = Cliente.objects.count()
    total_mascotas = Mascota.objects.count()
    total_usuarios = User.objects.count()

    # ESTADÍSTICAS DE ACTIVIDAD DE USUARIOS
    from django.contrib.sessions.models import Session
    import datetime

    # Usuarios activos (con sesiones activas en las últimas 24 horas)
    sesiones_activas = Session.objects.filter(expire_date__gte=timezone.now())
    usuarios_activos_24h = set()

    for session in sesiones_activas:
        try:
            session_data = session.get_decoded()
            user_id = session_data.get('_auth_user_id')
            if user_id:
                usuarios_activos_24h.add(int(user_id))
        except:
            continue

    usuarios_activos_count = len(usuarios_activos_24h)

    # Usuarios que iniciaron sesión en los últimos 7 días
    usuarios_recientes = User.objects.filter(
        last_login__gte=timezone.now() - datetime.timedelta(days=7)
    ).count()

    # Usuarios que nunca han iniciado sesión
    usuarios_sin_login = User.objects.filter(last_login__isnull=True).count()

    # Lista de usuarios recientes para mostrar en la tabla
    usuarios_recientes_list = User.objects.order_by('-last_login')[:10]
    
    # Citas
    citas_hoy = Cita.objects.filter(
        fecha__date=hoy,
        estado__in=['programada', 'confirmada']
    ).select_related('mascota__cliente__usuario').order_by('fecha')

    citas_mes = Cita.objects.filter(
        fecha__date__month=hoy.month,
        fecha__date__year=hoy.year
    ).count()

    citas_completadas_mes = Cita.objects.filter(
        fecha__date__month=hoy.month,
        fecha__date__year=hoy.year,
        estado='completada'
    ).count()
    
    # Clientes
    clientes_nuevos_mes = Cliente.objects.filter(
        fecha_registro__month=hoy.month,
        fecha_registro__year=hoy.year
    ).count()
    
    # Ingresos (asumiendo que tienes campo precio en Cita)
    try:
        ingresos_mes = Cita.objects.filter(
            fecha__date__month=hoy.month,
            fecha__date__year=hoy.year,
            estado='completada'
        ).aggregate(total=Sum('precio'))['total'] or 0
    except:
        ingresos_mes = 0
    
    # NUEVAS ESTADÍSTICAS DE LA TIENDA
    total_productos = Producto.objects.filter(activo=True).count()
    total_categorias = Categoria.objects.filter(activo=True).count()
    total_ordenes = Orden.objects.count()
    
    # Estadísticas de ventas de la tienda
    try:
        ventas_mes_tienda = Orden.objects.filter(
            fecha_creacion__month=hoy.month,
            fecha_creacion__year=hoy.year,
            estado__in=['entregada', 'enviada', 'en_proceso']
        ).aggregate(total=Sum('total'))['total'] or 0
    except:
        ventas_mes_tienda = 0
    
    ordenes_pendientes = Orden.objects.filter(estado='pendiente').count()
    productos_stock_bajo = Producto.objects.filter(stock__lte=F('stock_minimo'), activo=True).count()
    
    # Órdenes recientes (para el resumen)
    ordenes_recientes = Orden.objects.select_related('usuario').order_by('-fecha_creacion')[:5]

    # FILTROS PARA GESTIÓN DE ÓRDENES
    orden_fecha_inicio = request.GET.get('orden_fecha_inicio')
    orden_fecha_fin = request.GET.get('orden_fecha_fin')
    orden_estado = request.GET.get('orden_estado', 'todos')
    orden_busqueda = request.GET.get('orden_busqueda', '')

    # Convertir fechas de órdenes
    if orden_fecha_inicio:
        try:
            orden_fecha_inicio = datetime.strptime(orden_fecha_inicio, '%Y-%m-%d').date()
        except ValueError:
            orden_fecha_inicio = None
    if orden_fecha_fin:
        try:
            orden_fecha_fin = datetime.strptime(orden_fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            orden_fecha_fin = None

    # Base query para órdenes
    ordenes_query = Orden.objects.select_related('usuario')

    # Aplicar filtros de fecha
    if orden_fecha_inicio and orden_fecha_fin:
        ordenes_query = ordenes_query.filter(fecha_creacion__date__range=[orden_fecha_inicio, orden_fecha_fin])
    elif orden_fecha_inicio:
        ordenes_query = ordenes_query.filter(fecha_creacion__date__gte=orden_fecha_inicio)
    elif orden_fecha_fin:
        ordenes_query = ordenes_query.filter(fecha_creacion__date__lte=orden_fecha_fin)

    # Filtro por estado
    if orden_estado and orden_estado != 'todos':
        ordenes_query = ordenes_query.filter(estado=orden_estado)

    # Filtro por búsqueda
    if orden_busqueda:
        ordenes_query = ordenes_query.filter(
            Q(numero_orden__icontains=orden_busqueda) |
            Q(usuario__username__icontains=orden_busqueda) |
            Q(usuario__first_name__icontains=orden_busqueda) |
            Q(usuario__last_name__icontains=orden_busqueda)
        )

    # Órdenes para gestión (con filtros aplicados)
    ordenes_gestion = ordenes_query.order_by('-fecha_creacion')[:50]
    
    # Productos destacados (para el resumen)
    productos_destacados = Producto.objects.filter(destacado=True, activo=True)[:5]

    # FILTROS PARA GESTIÓN DE PRODUCTOS
    categoria_filtro = request.GET.get('categoria', '')
    estado_filtro = request.GET.get('estado', 'todos')
    stock_filtro = request.GET.get('stock', 'todos')
    busqueda = request.GET.get('busqueda', '')

    # Base query para productos
    productos_query = Producto.objects.select_related('categoria')

    # Aplicar filtros
    if categoria_filtro:
        productos_query = productos_query.filter(categoria_id=categoria_filtro)

    if estado_filtro == 'activos':
        productos_query = productos_query.filter(activo=True)
    elif estado_filtro == 'inactivos':
        productos_query = productos_query.filter(activo=False)

    if stock_filtro == 'bajo':
        productos_query = productos_query.filter(stock__lte=F('stock_minimo'))
    elif stock_filtro == 'sin_stock':
        productos_query = productos_query.filter(stock=0)

    if busqueda:
        productos_query = productos_query.filter(
            Q(nombre__icontains=busqueda) |
            Q(descripcion__icontains=busqueda) |
            Q(categoria__nombre__icontains=busqueda)
        )

    # Productos para gestión (con filtros aplicados)
    productos_gestion = productos_query[:50]  # Limitar a 50 para rendimiento
    
    # Productos con stock bajo
    productos_bajo_stock = Producto.objects.filter(
        stock__lte=F('stock_minimo'), 
        activo=True
    ).order_by('stock')[:5]
    
    # DATOS PARA GRÁFICOS EXISTENTES
    # Citas por estado
    citas_por_estado = list(Cita.objects.values('estado').annotate(
        total=Count('id')
    ).order_by('-total'))
    
    # Mascotas por tipo
    mascotas_por_tipo = list(Mascota.objects.values('tipo').annotate(
        total=Count('id')
    ).order_by('-total'))
    
    # Citas últimos 7 días
    citas_ultimos_7_dias = []
    for i in range(6, -1, -1):
        dia = hoy - timedelta(days=i)
        count = Cita.objects.filter(fecha__date=dia).count()
        citas_ultimos_7_dias.append({
            'dia': dia.strftime('%d/%m'),
            'total': count
        })
    
    # NUEVOS DATOS PARA GRÁFICOS DE TIENDA
    # Productos por categoría
    productos_por_categoria = list(Producto.objects.filter(activo=True).annotate(
        categoria_nombre=Coalesce('categoria__nombre', Value('SIN CATEGORÍA'))
    ).values('categoria_nombre').annotate(total=Count('id')).order_by('-total'))
    
    # Órdenes por estado
    ordenes_por_estado = list(Orden.objects.values('estado').annotate(
        total=Count('id')
    ).order_by('-total'))
    
    # Ventas últimos 7 días (tienda)
    ventas_ultimos_7_dias = []
    for i in range(6, -1, -1):
        dia = hoy - timedelta(days=i)
        ventas_dia = Orden.objects.filter(
            fecha_creacion__date=dia,
            estado__in=['entregada', 'enviada', 'en_proceso']
        ).aggregate(total=Sum('total'))['total'] or 0
        ventas_ultimos_7_dias.append({
            'dia': dia.strftime('%d/%m'),
            'total': float(ventas_dia)
        })
    
    # Próximas citas
    proximas_citas = Cita.objects.filter(
        fecha__date__gte=hoy,
        estado__in=['programada', 'confirmada']
    ).select_related('mascota__cliente__usuario').order_by('fecha')[:10]
    
    # Citas urgentes
    citas_urgentes = Cita.objects.filter(
        fecha__date=hoy,
        tipo='urgencia',
        estado__in=['programada', 'confirmada']
    ).count()

    # Cálculo de cumplimiento de citas
    cumplimiento_citas = 0
    if citas_mes > 0:
        cumplimiento_citas = round((citas_completadas_mes / citas_mes) * 100)

    # ESTADÍSTICAS DE NOTIFICACIONES
    total_notificaciones = Notificacion.objects.count()
    notificaciones_no_leidas = Notificacion.objects.filter(leida=False).count()
    notificaciones_ultima_semana = Notificacion.objects.filter(
        fecha_creacion__date__gte=hoy - timedelta(days=7)
    ).count()
    recordatorios_activos = Recordatorio.objects.filter(activo=True).count()
    recordatorios_pendientes = Recordatorio.objects.filter(
        activo=True,
        enviado=False,
        fecha_recordatorio__date__gte=hoy
    ).count()

    # FILTROS PARA GESTIÓN DE CITAS
    cita_busqueda = request.GET.get('cita_busqueda', '')
    cita_estado = request.GET.get('cita_estado', '')
    cita_tipo = request.GET.get('cita_tipo', '')
    cita_periodo = request.GET.get('cita_periodo', 'hoy')

    # Base query para citas
    citas_query = Cita.objects.select_related('mascota__cliente__usuario')

    # Aplicar filtros de período
    if cita_periodo == 'hoy':
        citas_query = citas_query.filter(fecha__date=hoy)
    elif cita_periodo == 'semana':
        semana_inicio = hoy - timedelta(days=hoy.weekday())
        semana_fin = semana_inicio + timedelta(days=6)
        citas_query = citas_query.filter(fecha__date__range=[semana_inicio, semana_fin])
    elif cita_periodo == 'mes':
        citas_query = citas_query.filter(fecha__date__month=hoy.month, fecha__date__year=hoy.year)
    # Para 'todas', no aplicamos filtro de fecha

    # Filtro por estado
    if cita_estado:
        citas_query = citas_query.filter(estado=cita_estado)

    # Filtro por tipo
    if cita_tipo:
        citas_query = citas_query.filter(tipo=cita_tipo)

    # Filtro por búsqueda
    if cita_busqueda:
        citas_query = citas_query.filter(
            Q(mascota__nombre__icontains=cita_busqueda) |
            Q(mascota__cliente__usuario__first_name__icontains=cita_busqueda) |
            Q(mascota__cliente__usuario__last_name__icontains=cita_busqueda) |
            Q(mascota__cliente__usuario__username__icontains=cita_busqueda) |
            Q(tipo__icontains=cita_busqueda)
        )

    # Citas filtradas para gestión (ordenadas por fecha)
    citas_hoy_filtradas = citas_query.order_by('fecha')[:50]

    context = {
        # ESTADÍSTICAS EXISTENTES
        'total_clientes': total_clientes,
        'total_mascotas': total_mascotas,
        'total_usuarios': total_usuarios,

        # ESTADÍSTICAS DE ACTIVIDAD DE USUARIOS
        'usuarios_activos_24h': usuarios_activos_count,
        'usuarios_recientes': usuarios_recientes,
        'usuarios_sin_login': usuarios_sin_login,
        'usuarios_recientes_list': usuarios_recientes_list,
        'citas_mes': citas_mes,
        'citas_completadas_mes': citas_completadas_mes,
        'clientes_nuevos_mes': clientes_nuevos_mes,
        'ingresos_mes': ingresos_mes,
        'citas_urgentes': citas_urgentes,
        'cumplimiento_citas': cumplimiento_citas,

        # NUEVAS ESTADÍSTICAS TIENDA
        'total_productos': total_productos,
        'total_categorias': total_categorias,
        'total_ordenes': total_ordenes,
        'ventas_mes_tienda': ventas_mes_tienda,
        'ordenes_pendientes': ordenes_pendientes,
        'productos_stock_bajo': productos_stock_bajo,
        'ordenes_recientes': ordenes_recientes,
        'productos_destacados': productos_destacados,
        'productos_bajo_stock': productos_bajo_stock,

        # LISTAS EXISTENTES
        'citas_hoy': citas_hoy_filtradas,
        'proximas_citas': proximas_citas,

        # FILTROS CITAS
        'cita_busqueda': cita_busqueda,
        'cita_estado': cita_estado,
        'cita_tipo': cita_tipo,
        'cita_periodo': cita_periodo,

        # DATOS PARA GRÁFICOS EXISTENTES
        'citas_por_estado_json': json.dumps(citas_por_estado, cls=DjangoJSONEncoder),
        'mascotas_por_tipo_json': json.dumps(mascotas_por_tipo, cls=DjangoJSONEncoder),
        'citas_ultimos_7_dias_json': json.dumps(citas_ultimos_7_dias, cls=DjangoJSONEncoder),

        # NUEVOS DATOS PARA GRÁFICOS DE TIENDA
        'productos_por_categoria_json': json.dumps(productos_por_categoria, cls=DjangoJSONEncoder),
        'ordenes_por_estado_json': json.dumps(ordenes_por_estado, cls=DjangoJSONEncoder),
        'ventas_ultimos_7_dias_json': json.dumps(ventas_ultimos_7_dias, cls=DjangoJSONEncoder),

        # FECHAS EXISTENTES
        'hoy': hoy,
        'mes_actual': hoy.strftime('%B %Y'),

        # FILTROS PRODUCTOS
        'productos_gestion': productos_gestion,
        'categorias': Categoria.objects.filter(activo=True),
        'categoria_filtro': categoria_filtro,
        'estado_filtro': estado_filtro,
        'stock_filtro': stock_filtro,
        'busqueda': busqueda,

        # FILTROS ÓRDENES
        'ordenes_gestion': ordenes_gestion,
        'orden_fecha_inicio': orden_fecha_inicio,
        'orden_fecha_fin': orden_fecha_fin,
        'orden_fecha_inicio_str': orden_fecha_inicio.strftime('%Y-%m-%d') if orden_fecha_inicio else '',
        'orden_fecha_fin_str': orden_fecha_fin.strftime('%Y-%m-%d') if orden_fecha_fin else '',
        'orden_estado': orden_estado,
        'orden_busqueda': orden_busqueda,
        'estados_orden': Orden.ESTADO_ORDEN,

        # ESTADÍSTICAS DE NOTIFICACIONES
        'total_notificaciones': total_notificaciones,
        'notificaciones_no_leidas': notificaciones_no_leidas,
        'notificaciones_ultima_semana': notificaciones_ultima_semana,
        'recordatorios_activos': recordatorios_activos,
        'recordatorios_pendientes': recordatorios_pendientes,
        'notificaciones_recientes': Notificacion.objects.select_related('cliente__usuario').order_by('-fecha_creacion')[:5],
    }
    
    return render(request, 'administracion/dashboard.html', context)

def exportar_datos(request):
    if request.method not in ['GET', 'POST']:
        return redirect('administracion:dashboard_admin')

    formato = request.POST.get('formato') or request.GET.get('formato') or 'excel'

    # Utilidades comunes
    hoy = timezone.now().date()
    generado_por = request.user.get_full_name() or request.user.username
    fecha_gen = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Construir datos
    stats_generales = [
        ['Métrica', 'Valor'],
        ['Total de Clientes', Cliente.objects.count()],
        ['Total de Mascotas', Mascota.objects.count()],
        ['Total de Usuarios', User.objects.count()],
        ['Total de Productos', Producto.objects.filter(activo=True).count()],
        ['Total de Categorías', Categoria.objects.filter(activo=True).count()],
        ['Total de Órdenes', Orden.objects.count()],
    ]
    stats_mes = [
        ['Métrica', 'Valor'],
        ['Citas este mes', Cita.objects.filter(fecha__date__month=hoy.month, fecha__date__year=hoy.year).count()],
        ['Citas completadas', Cita.objects.filter(fecha__date__month=hoy.month, fecha__date__year=hoy.year, estado='completada').count()],
        ['Clientes nuevos', Cliente.objects.filter(fecha_registro__month=hoy.month, fecha_registro__year=hoy.year).count()],
        ['Ventas tienda', Orden.objects.filter(fecha_creacion__month=hoy.month, fecha_creacion__year=hoy.year).aggregate(total=Sum('total'))['total'] or 0],
        ['Órdenes tienda', Orden.objects.filter(fecha_creacion__month=hoy.month, fecha_creacion__year=hoy.year).count()],
    ]
    citas_estado = [['Estado', 'Cantidad']] + [[e['estado'], e['total']] for e in Cita.objects.values('estado').annotate(total=Count('id'))]
    mascotas_tipo = [['Tipo', 'Cantidad']] + [[e['tipo'], e['total']] for e in Mascota.objects.values('tipo').annotate(total=Count('id'))]
    prod_por_categoria = [['Categoría', 'Cantidad']] + [[e['categoria__nombre'] or 'Sin categoría', e['total']] for e in Producto.objects.filter(activo=True).values('categoria__nombre').annotate(total=Count('id'))]
    ordenes_estado = [['Estado', 'Cantidad']] + [[e['estado'], e['total']] for e in Orden.objects.values('estado').annotate(total=Count('id'))]
    productos_bajo = [['Producto', 'Stock Actual', 'Stock Mínimo']]
    for p in Producto.objects.filter(stock__lte=F('stock_minimo'), activo=True).order_by('stock'):
        productos_bajo.append([p.nombre, p.stock, p.stock_minimo])
    citas_hoy_rows = [['Hora', 'Mascota', 'Dueño', 'Tipo', 'Estado']]
    for c in Cita.objects.filter(fecha__date=hoy).select_related('mascota__cliente__usuario'):
        citas_hoy_rows.append([
            c.fecha.strftime('%H:%M'),
            c.mascota.nombre,
            c.mascota.cliente.usuario.get_full_name() or c.mascota.cliente.usuario.username,
            c.get_tipo_display(),
            c.get_estado_display(),
        ])
    ordenes_pend_rows = [['Número Orden', 'Cliente', 'Total', 'Fecha']]
    for o in Orden.objects.filter(estado='pendiente').select_related('usuario'):
        ordenes_pend_rows.append([o.numero_orden, o.usuario.get_full_name() or o.usuario.username, float(o.total), o.fecha_creacion.strftime('%d/%m/%Y')])

    # Datos completos de la base de datos
    clientes_rows = [['ID', 'Usuario', 'Nombre', 'Apellido', 'Email', 'Teléfono', 'Dirección', 'Fecha Registro']]
    for c in Cliente.objects.select_related('usuario'):
        clientes_rows.append([
            c.id,
            c.usuario.username,
            c.usuario.first_name,
            c.usuario.last_name,
            c.usuario.email,
            c.telefono or '',
            c.direccion or '',
            c.fecha_registro.strftime('%d/%m/%Y')
        ])

    mascotas_rows = [['ID', 'Nombre', 'Tipo', 'Raza', 'Sexo', 'Fecha Nacimiento', 'Cliente', 'Fecha Registro']]
    for m in Mascota.objects.select_related('cliente__usuario'):
        mascotas_rows.append([
            m.id,
            m.nombre,
            m.get_tipo_display(),
            m.raza or '',
            m.get_sexo_display(),
            m.fecha_nacimiento.strftime('%d/%m/%Y') if m.fecha_nacimiento else '',
            m.cliente.usuario.get_full_name() or m.cliente.usuario.username,
            m.fecha_registro.strftime('%d/%m/%Y')
        ])

    citas_rows = [['ID', 'Fecha', 'Mascota', 'Cliente', 'Tipo', 'Estado', 'Notas']]
    for c in Cita.objects.select_related('mascota__cliente__usuario'):
        citas_rows.append([
            c.id,
            c.fecha.strftime('%d/%m/%Y %H:%M'),
            c.mascota.nombre,
            c.mascota.cliente.usuario.get_full_name() or c.mascota.cliente.usuario.username,
            c.get_tipo_display(),
            c.get_estado_display(),
            c.notas or ''
        ])

    productos_rows = [['ID', 'Nombre', 'Descripción', 'Categoría', 'Precio', 'Stock', 'Stock Mínimo', 'Activo', 'Destacado']]
    for p in Producto.objects.select_related('categoria'):
        productos_rows.append([
            p.id,
            p.nombre,
            p.descripcion or '',
            p.categoria.nombre if p.categoria else '',
            float(p.precio_final),
            p.stock,
            p.stock_minimo,
            'Sí' if p.activo else 'No',
            'Sí' if p.destacado else 'No'
        ])

    ordenes_rows = [['ID', 'Número Orden', 'Cliente', 'Subtotal', 'Impuesto', 'Envío', 'Total', 'Estado', 'Fecha Creación']]
    for o in Orden.objects.select_related('usuario'):
        ordenes_rows.append([
            o.id,
            o.numero_orden,
            o.usuario.get_full_name() or o.usuario.username,
            float(o.subtotal),
            float(o.impuesto),
            0,  # Envío no está en el modelo, asumir 0
            float(o.total),
            o.get_estado_display(),
            o.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        ])

    notificaciones_rows = [['ID', 'Cliente', 'Tipo', 'Título', 'Mensaje', 'Prioridad', 'Leída', 'Fecha Creación']]
    for n in Notificacion.objects.select_related('cliente__usuario'):
        notificaciones_rows.append([
            n.id,
            n.cliente.usuario.get_full_name() or n.cliente.usuario.username if n.cliente else 'Global',
            n.get_tipo_display(),
            n.titulo,
            n.mensaje[:100] + '...' if len(n.mensaje) > 100 else n.mensaje,
            n.get_prioridad_display(),
            'Sí' if n.leida else 'No',
            n.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        ])

    if formato.lower() == 'pdf':
        # Generación de PDF profesional con reportlab
        try:
            from reportlab.lib.pagesizes import A4, letter
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
            from reportlab.lib.units import inch
        except ImportError:
            messages.error(request, 'La exportación a PDF requiere reportlab. Instala con: pip install reportlab')
            return redirect('administracion:dashboard_admin')

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()

        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=20,
            alignment=1,  # Center
            textColor=colors.HexColor('#2c3e50')
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=15,
            textColor=colors.HexColor('#34495e')
        )

        section_style = ParagraphStyle(
            'CustomSection',
            parent=styles['Heading3'],
            fontSize=12,
            spaceAfter=10,
            textColor=colors.HexColor('#2c3e50'),
            borderColor=colors.HexColor('#3498db'),
            borderWidth=1,
            borderPadding=5,
            borderRadius=3
        )

        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=8
        )

        elems = []

        # Header con título principal
        title = Paragraph('<b>Reporte de Administración</b>', title_style)
        subtitle = Paragraph('<b>Veterinaria Vet Love</b>', subtitle_style)
        meta = Paragraph(f'<i>Generado por: {generado_por}<br/>Fecha: {fecha_gen}</i>', normal_style)

        elems += [title, subtitle, meta, Spacer(1, 20)]

        def add_professional_table(title_text, data, col_widths=None):
            # Título de sección
            elems.append(Paragraph(f'<b>{title_text}</b>', section_style))

            if not data:
                return

            # Crear tabla
            t = Table(data, colWidths=col_widths, hAlign='LEFT')

            # Estilos profesionales
            table_style = TableStyle([
                # Header
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 10),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),
                ('VALIGN', (0,0), (-1,0), 'MIDDLE'),
                ('BOTTOMPADDING', (0,0), (-1,0), 8),
                ('TOPPADDING', (0,0), (-1,0), 8),

                # Cuerpo
                ('BACKGROUND', (0,1), (-1,-1), colors.white),
                ('TEXTCOLOR', (0,1), (-1,-1), colors.black),
                ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
                ('FONTSIZE', (0,1), (-1,-1), 9),
                ('ALIGN', (0,1), (-1,-1), 'LEFT'),
                ('VALIGN', (0,1), (-1,-1), 'MIDDLE'),

                # Bordes
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#bdc3c7')),
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.HexColor('#ecf0f1')),

                # Filas alternas
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8f9fa')]),

                # Padding
                ('BOTTOMPADDING', (0,1), (-1,-1), 5),
                ('TOPPADDING', (0,1), (-1,-1), 5),
                ('LEFTPADDING', (0,0), (-1,-1), 6),
                ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ])

            t.setStyle(table_style)
            elems.extend([t, Spacer(1, 15)])

        # Primera página: Estadísticas principales
        add_professional_table('📊 Estadísticas Generales', stats_generales)
        add_professional_table('📈 Estadísticas del Mes Actual', stats_mes)

        # Salto de página
        elems.append(PageBreak())

        # Segunda página: Análisis detallado
        add_professional_table('📅 Citas por Estado', citas_estado)
        add_professional_table('🐾 Mascotas por Tipo', mascotas_tipo)
        add_professional_table('🛒 Productos por Categoría', prod_por_categoria)
        add_professional_table('📦 Órdenes por Estado', ordenes_estado)

        # Salto de página
        elems.append(PageBreak())

        # Tercera página: Alertas y pendientes
        add_professional_table('⚠️ Productos con Stock Bajo', productos_bajo)
        add_professional_table('📋 Citas para Hoy', citas_hoy_rows)
        add_professional_table('⏳ Órdenes Pendientes', ordenes_pend_rows)

        # Salto de página
        elems.append(PageBreak())

        # Cuarta página: Datos completos - Clientes
        add_professional_table('👥 Lista Completa de Clientes', clientes_rows)

        # Salto de página
        elems.append(PageBreak())

        # Quinta página: Mascotas
        add_professional_table('🐾 Lista Completa de Mascotas', mascotas_rows)

        # Salto de página
        elems.append(PageBreak())

        # Sexta página: Citas
        add_professional_table('📅 Lista Completa de Citas', citas_rows)

        # Salto de página
        elems.append(PageBreak())

        # Séptima página: Productos
        add_professional_table('🛒 Lista Completa de Productos', productos_rows)

        # Salto de página
        elems.append(PageBreak())

        # Octava página: Órdenes
        add_professional_table('📦 Lista Completa de Órdenes', ordenes_rows)

        # Salto de página
        elems.append(PageBreak())

        # Novena página: Notificaciones
        add_professional_table('🔔 Lista Completa de Notificaciones', notificaciones_rows)

        # Footer
        footer_text = f"Reporte generado el {date.today().strftime('%d/%m/%Y')} - Veterinaria Vet Love"
        footer = Paragraph(f'<i>{footer_text}</i>', ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray, alignment=1))
        elems.append(Spacer(1, 20))
        elems.append(footer)

        doc.build(elems)
        pdf_data = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_admin_{date.today().strftime('%Y%m%d')}.pdf"'
        response.write(pdf_data)
        return response
    
@login_required
@veterinario_required(login_url='/admin/login/')
def dashboard_veterinario(request):
    """
    Panel principal para veterinarios - vista especializada para gestión veterinaria
    """
    hoy = timezone.now().date()

    # Estadísticas específicas para veterinarios
    total_mascotas = Mascota.objects.count()
    citas_hoy = Cita.objects.filter(
        fecha__date=hoy,
        estado__in=['programada', 'confirmada']
    ).select_related('mascota__cliente__usuario').order_by('fecha')

    citas_pendientes = Cita.objects.filter(
        estado__in=['programada', 'confirmada']
    ).count()

    # Mascotas atendidas por el veterinario (si se implementa asignación)
    # Por ahora, mostrar todas las mascotas activas
    mascotas_activas = Mascota.objects.count()

    # Historial médico reciente
    try:
        historial_reciente = HistorialMedico.objects.select_related(
                'mascota__cliente__usuario'
            ).order_by('-fecha')[:10]
    except:
        historial_reciente = []

    # Vacunas próximas a vencer (próximos 30 días)
    try:
        proximas_vacunas = Vacuna.objects.filter(
                fecha_proxima__lte=hoy + timedelta(days=30),
                fecha_proxima__gte=hoy
            ).select_related('mascota__cliente__usuario').order_by('fecha_proxima')[:10]
    except:
        proximas_vacunas = []

    # Citas urgentes
    citas_urgentes = Cita.objects.filter(
            fecha__date=hoy,
            tipo='urgencia',
            estado__in=['programada', 'confirmada']
        ).count()

    # Estadísticas del mes
    citas_mes = Cita.objects.filter(
        fecha__date__month=hoy.month,
        fecha__date__year=hoy.year
    ).count()

    citas_completadas_mes = Cita.objects.filter(
        fecha__date__month=hoy.month,
        fecha__date__year=hoy.year,
        estado='completada'
    ).count()

    # Filtros para gestión de citas
    cita_busqueda = request.GET.get('cita_busqueda', '')
    cita_estado = request.GET.get('cita_estado', '')
    cita_tipo = request.GET.get('cita_tipo', '')
    cita_periodo = request.GET.get('cita_periodo', 'hoy')

    # Base query para citas
    citas_query = Cita.objects.select_related('mascota__cliente__usuario')

    # Aplicar filtros de período
    if cita_periodo == 'hoy':
        citas_query = citas_query.filter(fecha__date=hoy)
    elif cita_periodo == 'semana':
        semana_inicio = hoy - timedelta(days=hoy.weekday())
        semana_fin = semana_inicio + timedelta(days=6)
        citas_query = citas_query.filter(fecha__date__range=[semana_inicio, semana_fin])
    elif cita_periodo == 'mes':
        citas_query = citas_query.filter(fecha__date__month=hoy.month, fecha__date__year=hoy.year)

    # Filtro por estado
    if cita_estado:
        citas_query = citas_query.filter(estado=cita_estado)

    # Filtro por tipo
    if cita_tipo:
        citas_query = citas_query.filter(tipo=cita_tipo)

    # Filtro por búsqueda
    if cita_busqueda:
        citas_query = citas_query.filter(
            Q(mascota__nombre__icontains=cita_busqueda) |
            Q(mascota__cliente__usuario__first_name__icontains=cita_busqueda) |
            Q(mascota__cliente__usuario__last_name__icontains=cita_busqueda) |
            Q(mascota__cliente__usuario__username__icontains=cita_busqueda) |
            Q(tipo__icontains=cita_busqueda)
        )

    # Citas filtradas para gestión
    citas_gestion = citas_query.order_by('fecha')[:50]

    context = {
        # Estadísticas principales
        'total_mascotas': total_mascotas,
        'citas_hoy': citas_hoy,
        'citas_pendientes': citas_pendientes,
        'mascotas_activas': mascotas_activas,
        'historial_reciente': historial_reciente,
        'proximas_vacunas': proximas_vacunas,
        'citas_urgentes': citas_urgentes,
        'citas_mes': citas_mes,
        'citas_completadas_mes': citas_completadas_mes,

        # Filtros de citas
        'citas_gestion': citas_gestion,
        'cita_busqueda': cita_busqueda,
        'cita_estado': cita_estado,
        'cita_tipo': cita_tipo,
        'cita_periodo': cita_periodo,

        # Información del veterinario
        'veterinario': getattr(request.user, 'perfil_veterinario', None),
        'hoy': hoy,
        'mes_actual': hoy.strftime('%B %Y'),
    }

    return render(request, 'administracion/dashboard_veterinario.html', context)
    
    @login_required
    @veterinario_required(login_url='/admin/login/')
    def gestion_mascotas_veterinario(request):
        """
        Vista para que veterinarios gestionen mascotas
        """
        # Filtros
        tipo_filtro = request.GET.get('tipo', '')
        cliente_filtro = request.GET.get('cliente', '')
        busqueda = request.GET.get('busqueda', '')
        estado_filtro = request.GET.get('estado', 'activas')  # activas/inactivas/todas
    
        # Base query
        mascotas_query = Mascota.objects.select_related('cliente__usuario')
    
        # Aplicar filtros
        if tipo_filtro:
            mascotas_query = mascotas_query.filter(tipo=tipo_filtro)
    
        if cliente_filtro:
            mascotas_query = mascotas_query.filter(
                Q(cliente__usuario__username__icontains=cliente_filtro) |
                Q(cliente__usuario__first_name__icontains=cliente_filtro) |
                Q(cliente__usuario__last_name__icontains=cliente_filtro)
            )
    
        if estado_filtro == 'activas':
            mascotas_query = mascotas_query.filter(activo=True)
        elif estado_filtro == 'inactivas':
            mascotas_query = mascotas_query.filter(activo=False)
    
        if busqueda:
            mascotas_query = mascotas_query.filter(
                Q(nombre__icontains=busqueda) |
                Q(raza__icontains=busqueda) |
                Q(color__icontains=busqueda)
            )
    
        # Ordenar y limitar
        mascotas = mascotas_query.order_by('-fecha_registro')[:100]
    
        context = {
            'mascotas': mascotas,
            'tipo_filtro': tipo_filtro,
            'cliente_filtro': cliente_filtro,
            'busqueda': busqueda,
            'estado_filtro': estado_filtro,
            'tipos_mascota': Mascota.TIPO_CHOICES,
        }
    
        return render(request, 'administracion/gestion_mascotas_veterinario.html', context)
    
    @login_required
    @veterinario_required(login_url='/admin/login/')
    def gestion_citas_veterinario(request):
        """
        Vista para que veterinarios gestionen citas
        """
        hoy = timezone.now().date()
    
        # Filtros
        cita_busqueda = request.GET.get('cita_busqueda', '')
        cita_estado = request.GET.get('cita_estado', '')
        cita_tipo = request.GET.get('cita_tipo', '')
        cita_periodo = request.GET.get('cita_periodo', 'hoy')
    
        # Base query para citas
        citas_query = Cita.objects.select_related('mascota__cliente__usuario')
    
        # Aplicar filtros de período
        if cita_periodo == 'hoy':
            citas_query = citas_query.filter(fecha__date=hoy)
        elif cita_periodo == 'semana':
            semana_inicio = hoy - timedelta(days=hoy.weekday())
            semana_fin = semana_inicio + timedelta(days=6)
            citas_query = citas_query.filter(fecha__date__range=[semana_inicio, semana_fin])
        elif cita_periodo == 'mes':
            citas_query = citas_query.filter(fecha__date__month=hoy.month, fecha__date__year=hoy.year)
    
        # Filtro por estado
        if cita_estado:
            citas_query = citas_query.filter(estado=cita_estado)
    
        # Filtro por tipo
        if cita_tipo:
            citas_query = citas_query.filter(tipo=cita_tipo)
    
        # Filtro por búsqueda
        if cita_busqueda:
            citas_query = citas_query.filter(
                Q(mascota__nombre__icontains=cita_busqueda) |
                Q(mascota__cliente__usuario__first_name__icontains=cita_busqueda) |
                Q(mascota__cliente__usuario__last_name__icontains=cita_busqueda) |
                Q(mascota__cliente__usuario__username__icontains=cita_busqueda) |
                Q(tipo__icontains=cita_busqueda)
            )
    
        # Citas para gestión
        citas = citas_query.order_by('fecha')[:100]
    
        context = {
            'citas': citas,
            'cita_busqueda': cita_busqueda,
            'cita_estado': cita_estado,
            'cita_tipo': cita_tipo,
            'cita_periodo': cita_periodo,
            'hoy': hoy,
        }
    
        return render(request, 'administracion/gestion_citas_veterinario.html', context)
    
    @login_required
    @veterinario_required(login_url='/admin/login/')
    def historial_medico_veterinario(request, mascota_id):
        """
        Vista para que veterinarios vean y gestionen el historial médico de una mascota
        """
        try:
            mascota = Mascota.objects.select_related('cliente__usuario').get(id=mascota_id)
        except Mascota.DoesNotExist:
            messages.error(request, 'Mascota no encontrada.')
            return redirect('administracion:gestion_mascotas_veterinario')
    
        # Historial médico
        historial = HistorialMedico.objects.filter(mascota=mascota).order_by('-fecha')
    
        # Vacunas de la mascota
        vacunas = Vacuna.objects.filter(mascota=mascota).order_by('-fecha_aplicacion')
    
        # Citas de la mascota
        citas_mascota = Cita.objects.filter(mascota=mascota).order_by('-fecha')
    
        context = {
            'mascota': mascota,
            'historial': historial,
            'vacunas': vacunas,
            'citas_mascota': citas_mascota,
        }
    
        return render(request, 'administracion/historial_medico_veterinario.html', context)

    # Generación de Excel profesional con openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen'

    # Estilos
    header_fill = PatternFill('solid', fgColor='F1F3F5')
    header_font = Font(bold=True, color='212529')
    thin = Side(border_style='thin', color='DEE2E6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    def write_table(sheet, start_row, start_col, title, data):
        sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+len(data[0])-1)
        title_cell = sheet.cell(row=start_row, column=start_col, value=title)
        title_cell.font = Font(bold=True, size=12)
        row = start_row + 1
        # Header
        for j, val in enumerate(data[0], start=start_col):
            c = sheet.cell(row=row, column=j, value=val)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border
        # Rows
        for rdata in data[1:]:
            row += 1
            for j, val in enumerate(rdata, start=start_col):
                c = sheet.cell(row=row, column=j, value=val)
                c.border = border
        # Auto width
        for idx in range(start_col, start_col + len(data[0])):
            letter = get_column_letter(idx)
            max_len = 0
            for r in range(start_row, row+1):
                v = sheet.cell(row=r, column=idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            sheet.column_dimensions[letter].width = min(max_len + 2, 40)
        return row + 2

    # Header general
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Reporte de Administración - Veterinaria Vet Love'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = f'Generado por: {generado_por}   Fecha: {fecha_gen}'

    r = 4
    r = write_table(ws, r, 1, 'Estadísticas Generales', stats_generales)
    r = write_table(ws, r, 1, 'Estadísticas del Mes', stats_mes)

    # Hojas adicionales
    def add_sheet(name, data):
        sh = wb.create_sheet(title=name)
        write_table(sh, 1, 1, name, data)

    add_sheet('Citas por Estado', citas_estado)
    add_sheet('Mascotas por Tipo', mascotas_tipo)
    add_sheet('Prod. por Categoría', prod_por_categoria)
    add_sheet('Órdenes por Estado', ordenes_estado)
    add_sheet('Stock Bajo', productos_bajo)
    add_sheet('Citas Hoy', citas_hoy_rows)
    add_sheet('Órdenes Pendientes', ordenes_pend_rows)

    # Hojas con datos completos
    add_sheet('Todos los Clientes', clientes_rows)
    add_sheet('Todas las Mascotas', mascotas_rows)
    add_sheet('Todas las Citas', citas_rows)
    add_sheet('Todos los Productos', productos_rows)
    add_sheet('Todas las Órdenes', ordenes_rows)
    add_sheet('Todas las Notificaciones', notificaciones_rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_admin_{date.today().strftime('%Y%m%d')}.xlsx"'
    return response

def estadisticas_api(request):
    if request.method == 'GET':
        hoy = timezone.now().date()
        
        # Citas por día de la semana (últimas 4 semanas)
        datos_semana = []
        dias_semana = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
        
        for i in range(7):
            # Promedio de las últimas 4 semanas para este día
            citas_dia = Cita.objects.filter(
                fecha__week_day=i+2,  # Django: 2=Lunes, 3=Martes, etc.
                fecha__date__gte=hoy - timedelta(weeks=4)
            ).count()
            promedio = citas_dia / 4
            
            datos_semana.append({
                'dia': dias_semana[i],
                'promedio': round(promedio, 1)
            })
        
        # Citas últimos 30 días
        citas_30_dias = []
        for i in range(29, -1, -1):
            fecha = hoy - timedelta(days=i)
            count = Cita.objects.filter(fecha__date=fecha).count()
            citas_30_dias.append({
                'fecha': fecha.strftime('%d/%m'),
                'total': count
            })
        
        # Ventas últimos 30 días (tienda)
        ventas_30_dias = []
        for i in range(29, -1, -1):
            fecha = hoy - timedelta(days=i)
            ventas_dia = Orden.objects.filter(
                fecha_creacion__date=fecha,
                estado__in=['entregada', 'enviada', 'en_proceso']
            ).aggregate(total=Sum('total'))['total'] or 0
            ventas_30_dias.append({
                'fecha': fecha.strftime('%d/%m'),
                'total': float(ventas_dia)
            })
        
        return JsonResponse({
            'citas_semana': datos_semana,
            'citas_30_dias': citas_30_dias,
            'ventas_30_dias': ventas_30_dias,
            'mascotas_tipo': list(Mascota.objects.values('tipo').annotate(total=Count('id'))),
            'citas_estado': list(Cita.objects.values('estado').annotate(total=Count('id'))),
            'productos_categoria': list(Producto.objects.filter(activo=True).values('categoria__nombre').annotate(total=Count('id'))),
            'ordenes_estado': list(Orden.objects.values('estado').annotate(total=Count('id')))
        })

@login_required
@staff_required(login_url='/admin/login/')
def gestion_notificaciones(request):
    """
    Vista para que los administradores gestionen todas las notificaciones del sistema
    """
    # Filtros
    tipo_filtro = request.GET.get('tipo', '')
    cliente_filtro = request.GET.get('cliente', '')
    estado_filtro = request.GET.get('estado', 'todos')  # leida/no_leida/todas
    destinatario_filtro = request.GET.get('destinatario', 'todos')  # cliente/global/todos
    busqueda = request.GET.get('busqueda', '')

    # Base query
    notificaciones_query = Notificacion.objects.select_related('cliente__usuario')

    # Aplicar filtros
    if tipo_filtro:
        notificaciones_query = notificaciones_query.filter(tipo=tipo_filtro)

    if cliente_filtro:
        # Buscar en username, nombre, apellido, email
        notificaciones_query = notificaciones_query.filter(
            Q(cliente__usuario__username__icontains=cliente_filtro) |
            Q(cliente__usuario__first_name__icontains=cliente_filtro) |
            Q(cliente__usuario__last_name__icontains=cliente_filtro) |
            Q(cliente__usuario__email__icontains=cliente_filtro)
        )

    if estado_filtro == 'leida':
        notificaciones_query = notificaciones_query.filter(leida=True)
    elif estado_filtro == 'no_leida':
        notificaciones_query = notificaciones_query.filter(leida=False)

    # Filtro por destinatario
    if destinatario_filtro == 'cliente':
        notificaciones_query = notificaciones_query.filter(cliente__isnull=False)
    elif destinatario_filtro == 'global':
        notificaciones_query = notificaciones_query.filter(cliente__isnull=True)

    if busqueda:
        notificaciones_query = notificaciones_query.filter(
            Q(titulo__icontains=busqueda) |
            Q(mensaje__icontains=busqueda)
        )

    # Ordenar por fecha de creación descendente
    notificaciones = notificaciones_query.order_by('-fecha_creacion')[:100]  # Limitar a 100 para rendimiento

    # Estadísticas para las tarjetas
    notificaciones_no_leidas = Notificacion.objects.filter(leida=False).count()

    context = {
        'notificaciones': notificaciones,
        'tipo_filtro': tipo_filtro,
        'cliente_filtro': cliente_filtro,
        'estado_filtro': estado_filtro,
        'destinatario_filtro': destinatario_filtro,
        'busqueda': busqueda,
        'tipos_notificacion': Notificacion.TIPO_CHOICES,
        'notificaciones_no_leidas': notificaciones_no_leidas,
    }

    return render(request, 'administracion/gestion_notificaciones.html', context)

@login_required
@staff_required(login_url='/admin/login/')
def crear_notificacion(request):
    """
    Vista para que los administradores creen notificaciones personalizadas
    """
    if request.method == 'POST':
        from django.forms import ModelForm, ChoiceField

        class NotificacionForm(ModelForm):
            DESTINATARIO_CHOICES = [
                ('cliente', 'Cliente específico'),
                ('todos', 'Todos los usuarios'),
            ]

            destinatario = ChoiceField(
                choices=DESTINATARIO_CHOICES,
                initial='cliente',
                widget=forms.Select(attrs={'class': 'form-select'})
            )

            class Meta:
                model = Notificacion
                fields = ['cliente', 'tipo', 'titulo', 'mensaje', 'prioridad', 'url_relacionada']
                from django import forms
                cliente = forms.ModelChoiceField(
                    queryset=Cliente.objects.all(),
                    required=False,
                    empty_label="Seleccionar cliente",
                    widget=forms.Select(attrs={'class': 'form-select'})
                )
                tipo = forms.ChoiceField(
                    choices=Notificacion.TIPO_CHOICES,
                    widget=forms.Select(attrs={'class': 'form-select'})
                )
                prioridad = forms.ChoiceField(
                    choices=Notificacion.PRIORIDAD_CHOICES,
                    widget=forms.Select(attrs={'class': 'form-select'})
                )

            def clean(self):
                cleaned_data = super().clean()
                destinatario = cleaned_data.get('destinatario')
                cliente = cleaned_data.get('cliente')

                if destinatario == 'cliente' and not cliente:
                    raise forms.ValidationError("Debe seleccionar un cliente cuando el destinatario es específico.")
                elif destinatario == 'todos':
                    cleaned_data['cliente'] = None  # Forzar None para notificaciones globales

                return cleaned_data

        form = NotificacionForm(request.POST)
        if form.is_valid():
            destinatario = form.cleaned_data['destinatario']

            if destinatario == 'todos':
                # Crear notificación para todos los clientes
                clientes = Cliente.objects.all()
                notificaciones_creadas = 0

                for cliente in clientes:
                    Notificacion.objects.create(
                        cliente=cliente,
                        tipo=form.cleaned_data['tipo'],
                        titulo=form.cleaned_data['titulo'],
                        mensaje=form.cleaned_data['mensaje'],
                        prioridad=form.cleaned_data['prioridad'],
                        url_relacionada=form.cleaned_data.get('url_relacionada')
                    )
                    notificaciones_creadas += 1

                messages.success(request, f'Notificación enviada exitosamente a {notificaciones_creadas} clientes')
            else:
                # Crear notificación para cliente específico
                notificacion = form.save()
                cliente_nombre = notificacion.cliente.usuario.get_full_name() or notificacion.cliente.usuario.username
                messages.success(request, f'Notificación creada exitosamente para {cliente_nombre}')

            return redirect('administracion:gestion_notificaciones')
    else:
        from django.forms import ModelForm, ChoiceField

        class NotificacionForm(ModelForm):
            DESTINATARIO_CHOICES = [
                ('cliente', 'Cliente específico'),
                ('todos', 'Todos los usuarios'),
            ]

            destinatario = ChoiceField(
                choices=DESTINATARIO_CHOICES,
                initial='cliente',
                widget=forms.Select(attrs={'class': 'form-select'})
            )

            class Meta:
                model = Notificacion
                fields = ['cliente', 'tipo', 'titulo', 'mensaje', 'prioridad', 'url_relacionada']
                from django import forms
                cliente = forms.ModelChoiceField(
                    queryset=Cliente.objects.all(),
                    required=False,
                    empty_label="Seleccionar cliente",
                    widget=forms.Select(attrs={'class': 'form-select'})
                )
                tipo = forms.ChoiceField(
                    choices=Notificacion.TIPO_CHOICES,
                    widget=forms.Select(attrs={'class': 'form-select'})
                )
                prioridad = forms.ChoiceField(
                    choices=Notificacion.PRIORIDAD_CHOICES,
                    widget=forms.Select(attrs={'class': 'form-select'})
                )

        form = NotificacionForm()

    context = {
        'form': form,
        'titulo': 'Crear Notificación Personalizada',
    }

    return render(request, 'administracion/crear_notificacion.html', context)

@login_required
@login_required
@staff_required(login_url='/admin/login/')
def gestionar_usuarios(request):
    """
    Vista para que los administradores gestionen usuarios del sistema
    """
    # Filtros
    tipo_filtro = request.GET.get('tipo', 'todos')  # todos, staff, veterinarios, clientes
    busqueda = request.GET.get('busqueda', '')
    estado_filtro = request.GET.get('estado', 'todos')  # todos, activos, inactivos

    # Base query para usuarios
    usuarios_query = User.objects.select_related()

    # Aplicar filtros
    if tipo_filtro == 'staff':
        usuarios_query = usuarios_query.filter(is_staff=True)
    elif tipo_filtro == 'veterinarios':
        usuarios_query = usuarios_query.filter(groups__name='Veterinarios')
    elif tipo_filtro == 'clientes':
        usuarios_query = usuarios_query.filter(is_staff=False).exclude(groups__name='Veterinarios')

    if estado_filtro == 'activos':
        usuarios_query = usuarios_query.filter(is_active=True)
    elif estado_filtro == 'inactivos':
        usuarios_query = usuarios_query.filter(is_active=False)

    if busqueda:
        usuarios_query = usuarios_query.filter(
            Q(username__icontains=busqueda) |
            Q(first_name__icontains=busqueda) |
            Q(last_name__icontains=busqueda) |
            Q(email__icontains=busqueda)
        )

    # Obtener usuarios
    usuarios = usuarios_query.order_by('-date_joined')[:100]

    # Estadísticas
    total_usuarios = User.objects.count()
    usuarios_staff = User.objects.filter(is_staff=True).count()
    usuarios_veterinarios = User.objects.filter(groups__name='Veterinarios').count()
    usuarios_clientes = User.objects.filter(is_staff=False).exclude(groups__name='Veterinarios').count()
    usuarios_activos = User.objects.filter(is_active=True).count()
    usuarios_inactivos = User.objects.filter(is_active=False).count()

    context = {
        'usuarios': usuarios,
        'tipo_filtro': tipo_filtro,
        'busqueda': busqueda,
        'estado_filtro': estado_filtro,
        'total_usuarios': total_usuarios,
        'usuarios_staff': usuarios_staff,
        'usuarios_veterinarios': usuarios_veterinarios,
        'usuarios_clientes': usuarios_clientes,
        'usuarios_activos': usuarios_activos,
        'usuarios_inactivos': usuarios_inactivos,
    }

    return render(request, 'administracion/gestion_usuarios.html', context)

@login_required
@staff_required(login_url='/admin/login/')
def crear_usuario(request):
    """
    Vista para crear nuevos usuarios (staff, veterinarios, clientes)
    """
    if request.method == 'POST':
        from django.contrib.auth.forms import UserCreationForm
        from django import forms

        class UsuarioForm(forms.ModelForm):
            TIPO_USUARIO_CHOICES = [
                ('cliente', 'Cliente'),
                ('veterinario', 'Veterinario'),
                ('staff', 'Administrador'),
            ]

            tipo_usuario = forms.ChoiceField(
                choices=TIPO_USUARIO_CHOICES,
                widget=forms.Select(attrs={'class': 'form-select'}),
                label='Tipo de Usuario'
            )

            password1 = forms.CharField(
                widget=forms.PasswordInput(attrs={'class': 'form-control'}),
                label='Contraseña'
            )
            password2 = forms.CharField(
                widget=forms.PasswordInput(attrs={'class': 'form-control'}),
                label='Confirmar Contraseña'
            )

            class Meta:
                model = User
                fields = ['username', 'first_name', 'last_name', 'email']
                widgets = {
                    'username': forms.TextInput(attrs={'class': 'form-control'}),
                    'first_name': forms.TextInput(attrs={'class': 'form-control'}),
                    'last_name': forms.TextInput(attrs={'class': 'form-control'}),
                    'email': forms.EmailInput(attrs={'class': 'form-control'}),
                }

            def clean(self):
                cleaned_data = super().clean()
                password1 = cleaned_data.get('password1')
                password2 = cleaned_data.get('password2')

                if password1 and password2 and password1 != password2:
                    raise forms.ValidationError('Las contraseñas no coinciden')

                return cleaned_data

        form = UsuarioForm(request.POST)
        if form.is_valid():
            tipo_usuario = form.cleaned_data['tipo_usuario']
            password = form.cleaned_data['password1']

            # Crear usuario
            user = form.save(commit=False)
            user.set_password(password)
            user.save()

            # Asignar permisos según el tipo
            if tipo_usuario == 'staff':
                user.is_staff = True
                user.is_superuser = True
                user.save()
                messages.success(request, f'Administrador {user.get_full_name()} creado exitosamente.')
            elif tipo_usuario == 'veterinario':
                from administracion.models import Veterinario
                user.is_staff = False
                user.is_superuser = False
                user.save()

                # Crear perfil de veterinario
                Veterinario.objects.create(
                    usuario=user,
                    nombre_completo=user.get_full_name(),
                    especialidad='Medicina General',
                    telefono='',
                    activo=True
                )
                messages.success(request, f'Veterinario {user.get_full_name()} creado exitosamente.')
            else:  # cliente
                from clientes.models import Cliente
                user.is_staff = False
                user.is_superuser = False
                user.save()

                # Crear perfil de cliente
                Cliente.objects.create(
                    usuario=user,
                    telefono='',
                    direccion='',
                    preferencias_comunicacion='email'
                )
                messages.success(request, f'Cliente {user.get_full_name()} creado exitosamente.')

            return redirect('administracion:gestion_usuarios')
    else:
        from django import forms

        class UsuarioForm(forms.ModelForm):
            TIPO_USUARIO_CHOICES = [
                ('cliente', 'Cliente'),
                ('veterinario', 'Veterinario'),
                ('staff', 'Administrador'),
            ]

            tipo_usuario = forms.ChoiceField(
                choices=TIPO_USUARIO_CHOICES,
                widget=forms.Select(attrs={'class': 'form-select'}),
                label='Tipo de Usuario'
            )

            password1 = forms.CharField(
                widget=forms.PasswordInput(attrs={'class': 'form-control'}),
                label='Contraseña'
            )
            password2 = forms.CharField(
                widget=forms.PasswordInput(attrs={'class': 'form-control'}),
                label='Confirmar Contraseña'
            )

            class Meta:
                model = User
                fields = ['username', 'first_name', 'last_name', 'email']
                widgets = {
                    'username': forms.TextInput(attrs={'class': 'form-control'}),
                    'first_name': forms.TextInput(attrs={'class': 'form-control'}),
                    'last_name': forms.TextInput(attrs={'class': 'form-control'}),
                    'email': forms.EmailInput(attrs={'class': 'form-control'}),
                }

        form = UsuarioForm()

    context = {
        'form': form,
        'titulo': 'Crear Nuevo Usuario',
    }

    return render(request, 'administracion/crear_usuario.html', context)

@login_required
@staff_required(login_url='/admin/login/')
def editar_usuario(request, user_id):
    """
    Vista para editar usuarios existentes
    """
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        messages.error(request, 'Usuario no encontrado.')
        return redirect('administracion:gestion_usuarios')

    if request.method == 'POST':
        from django import forms

        class UsuarioEditForm(forms.ModelForm):
            TIPO_USUARIO_CHOICES = [
                ('cliente', 'Cliente'),
                ('veterinario', 'Veterinario'),
                ('staff', 'Administrador'),
            ]

            tipo_usuario = forms.ChoiceField(
                choices=TIPO_USUARIO_CHOICES,
                widget=forms.Select(attrs={'class': 'form-select'}),
                label='Tipo de Usuario'
            )

            cambiar_password = forms.BooleanField(
                required=False,
                widget=forms.CheckboxInput(attrs={'class': 'form-check-input'}),
                label='Cambiar contraseña'
            )

            password1 = forms.CharField(
                widget=forms.PasswordInput(attrs={'class': 'form-control'}),
                label='Nueva Contraseña',
                required=False
            )
            password2 = forms.CharField(
                widget=forms.PasswordInput(attrs={'class': 'form-control'}),
                label='Confirmar Nueva Contraseña',
                required=False
            )

            class Meta:
                model = User
                fields = ['username', 'first_name', 'last_name', 'email', 'is_active']
                widgets = {
                    'username': forms.TextInput(attrs={'class': 'form-control'}),
                    'first_name': forms.TextInput(attrs={'class': 'form-control'}),
                    'last_name': forms.TextInput(attrs={'class': 'form-control'}),
                    'email': forms.EmailInput(attrs={'class': 'form-control'}),
                    'is_active': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
                }

            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                # Determinar el tipo actual del usuario
                if self.instance.is_staff:
                    self.fields['tipo_usuario'].initial = 'staff'
                elif hasattr(self.instance, 'perfil_veterinario'):
                    self.fields['tipo_usuario'].initial = 'veterinario'
                else:
                    self.fields['tipo_usuario'].initial = 'cliente'

            def clean(self):
                cleaned_data = super().clean()
                cambiar_password = cleaned_data.get('cambiar_password')
                password1 = cleaned_data.get('password1')
                password2 = cleaned_data.get('password2')

                if cambiar_password:
                    if not password1 or not password2:
                        raise forms.ValidationError('Debe ingresar ambas contraseñas')
                    if password1 != password2:
                        raise forms.ValidationError('Las contraseñas no coinciden')

                return cleaned_data

        form = UsuarioEditForm(request.POST, instance=user)
        if form.is_valid():
            tipo_usuario = form.cleaned_data['tipo_usuario']
            cambiar_password = form.cleaned_data.get('cambiar_password')
            password1 = form.cleaned_data.get('password1')

            user = form.save(commit=False)

            # Actualizar permisos según el tipo
            if tipo_usuario == 'staff':
                user.is_staff = True
                user.is_superuser = True
                # Remover de veterinarios si estaba
                if hasattr(user, 'perfil_veterinario'):
                    user.perfil_veterinario.delete()
                user.groups.clear()
            elif tipo_usuario == 'veterinario':
                user.is_staff = False
                user.is_superuser = False
                # Crear perfil de veterinario si no existe
                from administracion.models import Veterinario
                if not hasattr(user, 'perfil_veterinario'):
                    Veterinario.objects.create(
                        usuario=user,
                        nombre_completo=user.get_full_name(),
                        especialidad='Medicina General',
                        telefono='',
                        activo=True
                    )
                user.groups.clear()
                from django.contrib.auth.models import Group
                grupo_veterinarios, created = Group.objects.get_or_create(name='Veterinarios')
                user.groups.add(grupo_veterinarios)
            else:  # cliente
                user.is_staff = False
                user.is_superuser = False
                # Remover de veterinarios si estaba
                if hasattr(user, 'perfil_veterinario'):
                    user.perfil_veterinario.delete()
                user.groups.clear()
                # Crear perfil de cliente si no existe
                from clientes.models import Cliente
                if not hasattr(user, 'cliente_profile'):
                    Cliente.objects.create(
                        usuario=user,
                        telefono='',
                        direccion='',
                        preferencias_comunicacion='email'
                    )

            # Cambiar contraseña si se solicita
            if cambiar_password and password1:
                user.set_password(password1)

            user.save()
            messages.success(request, f'Usuario {user.get_full_name()} actualizado exitosamente.')
            return redirect('administracion:gestion_usuarios')
    else:
        from django import forms

        class UsuarioEditForm(forms.ModelForm):
            TIPO_USUARIO_CHOICES = [
                ('cliente', 'Cliente'),
                ('veterinario', 'Veterinario'),
                ('staff', 'Administrador'),
            ]

            tipo_usuario = forms.ChoiceField(
                choices=TIPO_USUARIO_CHOICES,
                widget=forms.Select(attrs={'class': 'form-select'}),
                label='Tipo de Usuario'
            )

            cambiar_password = forms.BooleanField(
                required=False,
                widget=forms.CheckboxInput(attrs={'class': 'form-check-input'}),
                label='Cambiar contraseña'
            )

            password1 = forms.CharField(
                widget=forms.PasswordInput(attrs={'class': 'form-control'}),
                label='Nueva Contraseña',
                required=False
            )
            password2 = forms.CharField(
                widget=forms.PasswordInput(attrs={'class': 'form-control'}),
                label='Confirmar Nueva Contraseña',
                required=False
            )

            class Meta:
                model = User
                fields = ['username', 'first_name', 'last_name', 'email', 'is_active']
                widgets = {
                    'username': forms.TextInput(attrs={'class': 'form-control'}),
                    'first_name': forms.TextInput(attrs={'class': 'form-control'}),
                    'last_name': forms.TextInput(attrs={'class': 'form-control'}),
                    'email': forms.EmailInput(attrs={'class': 'form-control'}),
                    'is_active': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
                }

            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                # Determinar el tipo actual del usuario
                if self.instance.is_staff:
                    self.fields['tipo_usuario'].initial = 'staff'
                elif hasattr(self.instance, 'perfil_veterinario'):
                    self.fields['tipo_usuario'].initial = 'veterinario'
                else:
                    self.fields['tipo_usuario'].initial = 'cliente'

        form = UsuarioEditForm(instance=user)

    context = {
        'form': form,
        'usuario': user,
        'titulo': f'Editar Usuario: {user.get_full_name()}',
    }

    return render(request, 'administracion/editar_usuario.html', context)

@staff_required(login_url='/admin/login/')
def cambiar_estado_cita(request, cita_id):
    """
    Vista AJAX para cambiar el estado de una cita
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})

    try:
        from citas.models import Cita
        cita = Cita.objects.get(id=cita_id)
        nuevo_estado = request.POST.get('estado')

        if not nuevo_estado:
            return JsonResponse({'success': False, 'error': 'Estado no especificado'})

        # Validar que el estado sea válido
        estados_validos = ['programada', 'confirmada', 'completada', 'cancelada', 'reprogramada']
        if nuevo_estado not in estados_validos:
            return JsonResponse({'success': False, 'error': 'Estado no válido'})

        # Validar transiciones de estado
        if cita.estado == 'completada' and nuevo_estado != 'completada':
            return JsonResponse({'success': False, 'error': 'No se puede cambiar el estado de una cita completada'})

        if cita.estado == 'cancelada' and nuevo_estado != 'cancelada':
            return JsonResponse({'success': False, 'error': 'No se puede cambiar el estado de una cita cancelada'})

        # Cambiar el estado
        cita.estado = nuevo_estado
        cita.save()

        # Crear notificación para el cliente si es necesario
        if nuevo_estado in ['confirmada', 'completada', 'cancelada']:
            from notificaciones.models import Notificacion

            mensajes = {
                'confirmada': f'Su cita para {cita.mascota.nombre} ha sido confirmada.',
                'completada': f'Su cita para {cita.mascota.nombre} ha sido completada.',
                'cancelada': f'Su cita para {cita.mascota.nombre} ha sido cancelada.'
            }

            Notificacion.objects.create(
                cliente=cita.mascota.cliente,
                tipo='cita',
                titulo=f'Cita {nuevo_estado}: {cita.mascota.nombre}',
                mensaje=mensajes.get(nuevo_estado, f'Su cita ha cambiado a {nuevo_estado}'),
                url_relacionada=f'/citas/mis-citas/'
            )

        return JsonResponse({
            'success': True,
            'nuevo_estado': nuevo_estado,
            'mensaje': f'Cita cambiada a {nuevo_estado} exitosamente'
        })

    except Cita.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cita no encontrada'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@staff_required(login_url='/admin/login/')
def preview_pagina_web(request):
    """
    Vista para que los administradores puedan ver cómo se ve la página web
    desde la perspectiva de un usuario no logueado
    """
    from tienda.models import Producto, Categoria

    # Obtener productos destacados para mostrar en la página de inicio
    productos_destacados = Producto.objects.filter(activo=True, destacado=True)[:6]
    categorias = Categoria.objects.filter(activo=True)

    context = {
        'productos_destacados': productos_destacados,
        'categorias': categorias,
        'preview_mode': True,  # Indicador de que estamos en modo preview
    }

    return render(request, 'inicio.html', context)

@login_required
@staff_required(login_url='/admin/login/')
def preview_tienda(request):
    """
    Vista para que los administradores puedan ver cómo se ve la tienda
    desde la perspectiva de un usuario normal
    """
    from tienda.models import Producto, Categoria

    # Obtener todas las categorías activas
    categorias = Categoria.objects.filter(activo=True)

    # Obtener productos activos con filtros aplicados
    productos = Producto.objects.filter(activo=True)

    # Filtros
    categoria_id = request.GET.get('categoria')
    tipo = request.GET.get('tipo')
    tipo_mascota = request.GET.get('tipo_mascota')
    buscar = request.GET.get('buscar')

    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    if tipo:
        productos = productos.filter(tipo=tipo)
    if tipo_mascota and tipo_mascota != 'todos':
        productos = productos.filter(tipo_mascota=tipo_mascota)
    if buscar:
        productos = productos.filter(nombre__icontains=buscar)

    context = {
        'productos': productos,
        'categorias': categorias,
        'categoria_actual': int(categoria_id) if categoria_id else None,
        'preview_mode': True,  # Indicador de que estamos en modo preview
    }

    return render(request, 'tienda/lista_productos.html', context)

@login_required
@staff_required(login_url='/admin/login/')
def descargar_plantillas_carnets(request):
    """
    Vista para seleccionar mascota y descargar plantilla individual de carnet
    """
    from mascotas.models import Mascota

    # Si es POST, procesar la selección de mascota
    if request.method == 'POST':
        mascota_id = request.POST.get('mascota_id')
        if not mascota_id:
            messages.error(request, 'Debe seleccionar una mascota.')
            return redirect('administracion:descargar_plantillas_carnets')

        try:
            mascota = Mascota.objects.select_related('cliente__usuario').get(id=mascota_id)
        except Mascota.DoesNotExist:
            messages.error(request, 'Mascota no encontrada.')
            return redirect('administracion:descargar_plantillas_carnets')

        # Verificar permisos
        es_admin = request.user.is_staff or request.user.is_superuser
        if not es_admin:
            messages.error(request, 'No tiene permisos para descargar plantillas.')
            return redirect('administracion:dashboard_admin')

        # Generar PDF de plantilla individual
        return generar_plantilla_carnet_pdf(mascota, request.user)

    # Si es GET, mostrar formulario de selección
    # Filtros para la lista de mascotas
    tipo_filtro = request.GET.get('tipo', 'todas')
    cliente_filtro = request.GET.get('cliente', '')
    busqueda = request.GET.get('busqueda', '')

    # Base query para mascotas
    mascotas_query = Mascota.objects.select_related('cliente__usuario').order_by('nombre')

    # Aplicar filtros
    if tipo_filtro and tipo_filtro != 'todas':
        mascotas_query = mascotas_query.filter(tipo=tipo_filtro)

    if cliente_filtro:
        mascotas_query = mascotas_query.filter(
            Q(cliente__usuario__username__icontains=cliente_filtro) |
            Q(cliente__usuario__first_name__icontains=cliente_filtro) |
            Q(cliente__usuario__last_name__icontains=cliente_filtro)
        )

    if busqueda:
        mascotas_query = mascotas_query.filter(
            Q(nombre__icontains=busqueda) |
            Q(raza__icontains=busqueda)
        )

    mascotas = mascotas_query[:100]  # Limitar para rendimiento

    context = {
        'mascotas': mascotas,
        'tipo_filtro': tipo_filtro,
        'cliente_filtro': cliente_filtro,
        'busqueda': busqueda,
        'tipos_mascota': Mascota.TIPO_CHOICES,
    }

    return render(request, 'administracion/plantillas_carnets.html', context)


def generar_plantilla_carnet_pdf(mascota, usuario):
    """
    Función auxiliar para generar PDF de plantilla de carnet individual
    """
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.units import inch
    from io import BytesIO
    from django.utils import timezone
    from datetime import timedelta

    # Crear buffer para el PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                           leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    width, height = A4
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=20,
        alignment=1,  # Centrado
        textColor=colors.black
    )

    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=10,
        textColor=colors.black
    )

    # Título principal centrado
    title = Paragraph('<b>CARNET DE VACUNACIÓN</b>', ParagraphStyle('CenteredTitle', parent=title_style, alignment=1))
    elements.append(title)

    # Subtítulo con nombre de la veterinaria
    subtitle = Paragraph('<b>VETERINARIA VET LOVE</b>', ParagraphStyle('CenteredSubtitle', parent=styles['Heading3'], fontSize=12, alignment=1, textColor=colors.HexColor('#2c3e50')))
    elements.append(subtitle)
    elements.append(Spacer(1, 5))

    # Información relevante adicional
    info_relevante = Paragraph(
        f"<b>Mascota:</b> {mascota.nombre.upper()} | <b>ID:</b> {mascota.id:08d} | <b>Generado:</b> {timezone.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle('InfoRelevante', parent=styles['Normal'], fontSize=8, alignment=1, textColor=colors.grey)
    )
    elements.append(info_relevante)
    elements.append(Spacer(1, 5))

    # Información de la mascota - COMPACTA
    elements.append(Spacer(1, 10))

    # Crear tabla compacta de información de la mascota
    mascota_data = [
        ['Nombre:', mascota.nombre.upper(), 'Especie:', mascota.get_tipo_display().upper()],
        ['Raza:', (mascota.raza or 'MESTIZO').upper(), 'Sexo:', mascota.get_sexo_display().upper()],
        ['Color:', (mascota.color or 'NO ESPECIFICADO').upper(), 'F. Nacimiento:', mascota.fecha_nacimiento.strftime('%d/%m/%Y') if mascota.fecha_nacimiento else 'NO REGISTRA'],
    ]

    mascota_table = Table(mascota_data, colWidths=[60, 120, 60, 120])
    mascota_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1), True),  # Permitir ajuste de palabras
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alineación vertical al centro
    ]))
    elements.append(mascota_table)

    # Información del propietario - COMPACTA
    elements.append(Spacer(1, 8))

    propietario_data = [
        ['Propietario:', (mascota.cliente.usuario.get_full_name() or mascota.cliente.usuario.username).upper(), 'Documento:', ''],
        ['Dirección:', (mascota.cliente.direccion or '').upper()[:25], 'Teléfono:', (mascota.cliente.telefono or '')],
    ]

    propietario_table = Table(propietario_data, colWidths=[60, 120, 60, 120])
    propietario_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1), True),  # Permitir ajuste de palabras
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alineación vertical al centro
    ]))
    elements.append(propietario_table)

    # Tabla de vacunas VACÍA para llenado manual - OPTIMIZADA PARA UNA PÁGINA
    elements.append(Spacer(1, 10))

    # Crear párrafo centrado para el título de la tabla
    titulo_centrado = Paragraph("REGISTRO DE VACUNACIÓN", section_style)
    titulo_centrado.style = ParagraphStyle(
        'CenteredSection',
        parent=section_style,
        alignment=1,  # Centrado
    )
    elements.append(titulo_centrado)

    # Encabezados de la tabla de vacunas - MÁS COMPACTOS
    vaccine_headers = ['Vacuna', 'F. Aplicación', 'Próxima', 'Lote', 'Lab.', 'Estado']

    # SOLO filas vacías para llenado manual (12 filas para mejor aprovechamiento del espacio con celdas más grandes)
    vaccine_data = [vaccine_headers]

    for j in range(12):
        vaccine_data.append([
            '',
            '',
            '',
            '',
            '',
            '○ Aplicada ○ Pendiente'
        ])

    # Crear tabla de vacunas - MÁS GRANDE PARA MEJOR VISIBILIDAD
    vaccine_table = Table(vaccine_data, colWidths=[90, 65, 65, 60, 60, 110])

    # Estilo de tabla Excel-like optimizado para llenado manual
    vaccine_table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.8, 0.8, 0.8)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 6),

        # Filas vacías para llenado manual
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 6),

        # Bordes más gruesos para fácil identificación
        ('GRID', (0, 0), (-1, -1), 1.2, colors.black),

        # Líneas internas más visibles
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.grey),
    ]))

    elements.append(vaccine_table)

    # Instrucciones y firma - MÁS COMPACTAS
    elements.append(Spacer(1, 8))

    instrucciones_style = ParagraphStyle(
        'Instrucciones',
        parent=styles['Normal'],
        fontSize=7,
        textColor=colors.black,
        spaceAfter=8
    )

    instrucciones = """
    <b>INSTRUCCIONES:</b> Complete cada fila con bolígrafo negro/azul. Marque ✓ en "Estado". Documento oficial Vet Love.
    """
    elements.append(Paragraph(instrucciones, instrucciones_style))

    # Espacio para firma y sello - MÁS COMPACTO
    firma_data = [
        ['Firma Veterinario:', '', 'Sello:', ''],
        ['Fecha:', '', 'Reg. Vet:', ''],
    ]

    firma_table = Table(firma_data, colWidths=[100, 180, 70, 120])
    firma_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 18),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('WORDWRAP', (0, 0), (-1, -1), True),  # Permitir ajuste de palabras
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alineación vertical al centro
    ]))
    elements.append(firma_table)

    # Pie de página
    elements.append(Spacer(1, 5))
    footer_text = f"PLANTILLA OFICIAL - ID: {mascota.id:08d} - Generada: {timezone.now().strftime('%d/%m/%Y %H:%M')} - Por: {usuario.get_full_name() or usuario.username}"
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=6,
        textColor=colors.grey,
        alignment=1
    )
    elements.append(Paragraph(footer_text, footer_style))

    # Generar PDF
    doc.build(elements)

    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Plantilla_Carnet_{mascota.nombre.replace(' ', '_')}.pdf"'

    return response